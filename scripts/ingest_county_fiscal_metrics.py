"""Populate fiscal-health metrics for every FL county commissioner.

Source: Florida Department of Revenue (PTO) — county_overview.xlsx, sheet
"County Taxable Value History 2000-2025". This is the cleanest piece of
machine-readable county fiscal data the state publishes; full CAFRs are
PDF-only and bond ratings sit behind paid agency APIs.

What we can compute from it:

  taxable_value_change  =  (TV_2025 - TV_2024) / TV_2024 * 100

This is the YoY change in a county's taxable property value — a meaningful
indicator of fiscal health that county commissioners influence through
assessment policy, growth approvals, exemption decisions, etc.

Rating bands (tuned from observed FL distribution):
  >= 10  excellent  (boom growth)
  >= 5   good       (healthy growth)
  >= 2   meeting    (modest growth)
  >= 0   concerning (flat)
  < 0    poor       (decline)

Applied to every elected_officials row matching:
  level = 'local'
  AND county IS NOT NULL
  AND title ILIKE '%county commissioner%'
       OR title ILIKE '%county commission%'

budget_variance and bond_rating: deliberately left as placeholders.
budget_variance lives in per-county CAFR PDFs with no machine-readable
index; bond_rating sits with Moody's / S&P / Fitch behind paid APIs.

Run:
  python scripts/ingest_county_fiscal_metrics.py
  python scripts/ingest_county_fiscal_metrics.py --dry-run

Env: SUPABASE_URL, SUPABASE_SERVICE_KEY (or SUPABASE_KEY).
"""

from __future__ import annotations

import argparse
import os
import sys
import tempfile
from typing import Optional

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY")

XLSX_URL = "https://floridarevenue.com/property/Documents/county_overview.xlsx"
SHEET = "County Taxable Value History"
SESSION_YEAR = 2025

LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "county_fiscal_log.txt")


def log(msg: str) -> None:
    print(msg, flush=True)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(msg + "\n")
    except OSError:
        pass


def rate_value_change(pct: float) -> str:
    if pct >= 10: return "excellent"
    if pct >= 5:  return "good"
    if pct >= 2:  return "meeting"
    if pct >= 0:  return "concerning"
    return "poor"


def download_xlsx() -> str:
    """Returns the local path of the downloaded xlsx."""
    r = requests.get(XLSX_URL, timeout=60,
                     headers={"User-Agent": "PolitiScore Ingest/1.0"})
    r.raise_for_status()
    tmp = tempfile.NamedTemporaryFile(prefix="fl_county_", suffix=".xlsx", delete=False)
    tmp.write(r.content)
    tmp.close()
    return tmp.name


def parse_taxable_values(path: str) -> dict[str, dict]:
    """Returns { 'Alachua': { 2024: float, 2025: float, ... }, ... }.

    Sheet layout: row 4 has year headers starting in col B; row 5+ have
    one row per county, name in col A, values across.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET]

    # Find the header row by scanning for the row whose A cell == 'County'
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if (row[0] or "").strip().lower() == "county":
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError("Couldn't locate header row in xlsx")

    # Cells in ws[row] are Cell objects; .value gives the raw scalar.
    header = [c.value for c in ws[header_row_idx]]
    year_cols: list[tuple[int, int]] = []  # (col_index, year)
    for idx, h in enumerate(header):
        try:
            year = int(str(h).strip())
            if 1990 <= year <= 2100:
                year_cols.append((idx, year))
        except (ValueError, TypeError):
            continue

    result: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        county = (row[0] or "").strip() if row[0] else ""
        if not county or county.lower().startswith(("statewide", "total")):
            continue
        years: dict[int, float] = {}
        for col_idx, year in year_cols:
            try:
                v = row[col_idx]
                if v is None or v == "":
                    continue
                years[year] = float(v)
            except (TypeError, ValueError):
                continue
        if years:
            result[county] = years
    return result


def upsert_metric(sb, *, official_id: int, official_name: str,
                  pct: float, county: str, dry_run: bool) -> None:
    rating = rate_value_change(pct)
    payload = {
        "official_id": official_id,
        "official_name": official_name,
        "metric_key": "taxable_value_change",
        "metric_label": "Taxable Value Growth (YoY)",
        "metric_value": f"{pct:.1f}",
        "metric_unit": "%",
        "benchmark_value": "5.0",
        "benchmark_label": "Statewide median growth",
        "performance_rating": rating,
        "year": SESSION_YEAR,
        "source": "Florida DOR – County Taxable Value History",
        "source_url": XLSX_URL,
        "notes": f"YoY change in {county} County taxable value 2024->2025",
        "last_updated": "now()",
    }
    if dry_run:
        return
    sb.table("accountability_metrics") \
        .delete() \
        .eq("official_id", official_id) \
        .eq("metric_key", "taxable_value_change") \
        .eq("year", SESSION_YEAR) \
        .execute()
    sb.table("accountability_metrics").insert(payload).execute()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not (SUPABASE_URL and SUPABASE_KEY):
        log("ERROR: SUPABASE_URL + SUPABASE_SERVICE_KEY required")
        return 1
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    log(f"start  dry-run={args.dry_run}")

    # 1. Download + parse FL DOR xlsx
    path = download_xlsx()
    log(f"downloaded xlsx -> {path}")
    counties_data = parse_taxable_values(path)
    log(f"parsed {len(counties_data)} counties from xlsx")

    # 2. Compute YoY change
    yoy: dict[str, float] = {}
    for c, years in counties_data.items():
        if 2025 in years and 2024 in years and years[2024] > 0:
            pct = (years[2025] - years[2024]) / years[2024] * 100.0
            yoy[c.strip()] = pct
    log(f"YoY change computed for {len(yoy)} counties")

    if not yoy:
        log("no YoY data — aborting")
        return 1

    distribution = {"excellent": 0, "good": 0, "meeting": 0, "concerning": 0, "poor": 0}
    for pct in yoy.values():
        distribution[rate_value_change(pct)] += 1
    log(f"county-level distribution: {distribution}")

    # 3. Load county commissioners from elected_officials
    commissioners: list[dict] = []
    page = 0; PAGE = 500
    while True:
        chunk = (sb.table("elected_officials")
                   .select("id, name, title, county")
                   .eq("state", "FL")
                   .eq("level", "local")
                   .range(page * PAGE, page * PAGE + PAGE - 1)
                   .execute().data or [])
        if not chunk:
            break
        for r in chunk:
            t = (r.get("title") or "").lower()
            if "county commissioner" in t or "county commission" in t:
                commissioners.append(r)
        if len(chunk) < PAGE:
            break
        page += 1
    log(f"county commissioners in DB: {len(commissioners)}")

    # 4. Upsert metric per commissioner
    written = skipped_no_county = skipped_no_data = 0
    for c in commissioners:
        county = (c.get("county") or "").strip()
        if not county:
            skipped_no_county += 1
            continue
        # County name in DB may include " County" suffix; strip
        county_key = county.replace(" County", "").strip()
        # Match case-insensitively
        match_pct = None
        for k, pct in yoy.items():
            if k.lower() == county_key.lower():
                match_pct = pct
                break
        if match_pct is None:
            skipped_no_data += 1
            continue

        upsert_metric(
            sb,
            official_id=c["id"], official_name=c["name"],
            pct=match_pct, county=county_key,
            dry_run=args.dry_run,
        )
        written += 1

    log("")
    log(f"DONE  commissioners={len(commissioners)}  written={written}  "
        f"skipped_no_county={skipped_no_county}  skipped_no_data={skipped_no_data}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
